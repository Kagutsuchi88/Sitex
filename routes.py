# routes.py - COMPLETO CON CONFIRMACIÓN DE EMAIL
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import current_user, login_user, logout_user, login_required
from flask_mail import Message
from werkzeug.utils import secure_filename
from datetime import date, datetime, timedelta
from sqlalchemy import func, extract
import os
import secrets
import pandas as pd

from extensions import db, mail
from models import Empleado, Tanque, Descargue, RegistroMedida, MedicionCargue, SesionActiva, Auditoria, Venta
from forms import (LoginForm, RegisterForm, MedicionForm, DescargueForm, ChangePasswordForm, 
                  ResetPasswordForm, RequestPasswordResetForm, PasswordResetForm, TanqueForm,
                  CargaMasivaForm, FiltroMedicionesForm)
from utils import (islero_or_encargado_required, admin_or_encargado_required, admin_required,
                  registrar_auditoria, allowed_file)

# Blueprints
auth_bp = Blueprint("auth", __name__, url_prefix="/auth")
main_bp = Blueprint("main", __name__)
dashboard_bp = Blueprint("dashboard", __name__, url_prefix="/dashboard")
medicion_bp = Blueprint("medicion", __name__, url_prefix="/medicion")
admin_bp = Blueprint("admin", __name__, url_prefix="/admin")

# ============= FUNCIONES AUXILIARES =============
def enviar_email_confirmacion(empleado, token):
    """Enviar email de confirmación al registrarse"""
    confirm_url = url_for('auth.confirm_email', token=token, _external=True)
    contrasena_temp = empleado.numero_documento[-4:] if len(empleado.numero_documento) >= 4 else empleado.numero_documento
    
    msg = Message("Confirma tu email - Hayuelos",
                recipients=[empleado.email])
    msg.body = f"""Hola {empleado.nombre_empleado} {empleado.apellido_empleado},

¡Bienvenido a Sitex!

Para completar tu registro, confirma tu email haciendo clic en el siguiente enlace:

{confirm_url}

Este enlace expira en 24 horas.

Tu usuario es: {empleado.usuario}
Tu contraseña temporal es: {contrasena_temp}

Por seguridad, te recomendamos cambiar tu contraseña después de iniciar sesión.

Si no te registraste en Hayuelos, ignora este correo.

Saludos,
Sistema Hayuelos - "Y También vendemos combustible"
"""
    msg.html = f"""
    <html>
    <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
        <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
            <h2 style="color: #E10000;">¡Bienvenido a Hayuelos!</h2>
            <p>Hola <strong>{empleado.nombre_empleado} {empleado.apellido_empleado}</strong>,</p>
            <p>Gracias por registrarte en el sistema Hayuelos. Para completar tu registro, confirma tu email:</p>
            
            <div style="text-align: center; margin: 30px 0;">
                <a href="{confirm_url}" 
                   style="background-color: #E10000; color: white; padding: 12px 30px; text-decoration: none; border-radius: 5px; display: inline-block;">
                    Confirmar Email
                </a>
            </div>
            
            <div style="background-color: #f8f9fa; padding: 15px; border-radius: 5px; margin: 20px 0;">
                <p><strong>Tus credenciales:</strong></p>
                <p>Usuario: <code>{empleado.usuario}</code></p>
                <p>Contraseña temporal: <code>{contrasena_temp}</code></p>
            </div>
            
            <p style="color: #666; font-size: 0.9em;">Este enlace expira en 24 horas.</p>
            <p style="color: #666; font-size: 0.9em;">Si no te registraste en Hayuelos, ignora este correo.</p>
            
            <hr style="margin: 30px 0; border: none; border-top: 1px solid #ddd;">
            <p style="color: #999; font-size: 0.8em; text-align: center;">
                Sistema Hayuelos - "Y También vendemos combustible"
            </p>
        </div>
    </body>
    </html>
    """
    mail.send(msg)

def calcular_altura_maxima(capacidad_galones):
    """Calcular altura máxima en cm basada en capacidad del tanque"""
    # Radio estándar en cm (ajustar según tanques reales)
    radio_cm = 125  # 2.5m de diámetro
    
    # Volumen en cm³ = capacidad en galones * 3785.411784
    volumen_cm3 = capacidad_galones * 3785.411784
    
    # Altura = Volumen / (π * r²)
    area_base = 3.14159 * (radio_cm ** 2)
    altura_cm = volumen_cm3 / area_base
    
    return round(altura_cm, 2)

# ============= AUTH ROUTES =============
@auth_bp.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard.index"))

    form = LoginForm()
    if form.validate_on_submit():
        usuario = form.username.data.strip()
        contrasena = form.password.data

        empleado = Empleado.query.filter(
            (Empleado.usuario == usuario) | (Empleado.numero_documento == usuario)
        ).first()

        if empleado and empleado.check_password(contrasena):
            if not empleado.activo:
                flash("Su cuenta ha sido deshabilitada. Contacte al administrador.", "danger")
                return redirect(url_for("auth.login"))

            # NUEVO: Verificar si el email está confirmado
            if not empleado.email_confirmado:
                flash("Debe confirmar su email antes de iniciar sesión. Revise su correo electrónico.", "warning")
                return redirect(url_for("auth.resend_confirmation"))

            login_user(empleado, remember=form.remember_me.data)
            
            # Crear sesión activa
            session_id = secrets.token_urlsafe(32)
            sesion = SesionActiva(
                id_empleados=empleado.id_empleados,
                session_id=session_id,
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent', '')[:255]
            )
            db.session.add(sesion)
            db.session.commit()

            flash(f"Bienvenido {empleado.nombre_empleado}!", "success")
            return redirect(url_for("dashboard.index"))
        else:
            flash("Usuario o contraseña incorrectos", "danger")

    return render_template("auth/login.html", form=form)

@auth_bp.route("/register", methods=["GET", "POST"])
def register():
    form = RegisterForm()
    if form.validate_on_submit():
        # Verificar duplicados
        if Empleado.query.filter_by(numero_documento=form.numero_documento.data).first():
            flash("El número de documento ya está registrado", "danger")
            return render_template("auth/register.html", form=form)
        
        if Empleado.query.filter_by(email=form.email.data).first():
            flash("El email ya está registrado", "danger")
            return render_template("auth/register.html", form=form)
        
        if Empleado.query.filter_by(usuario=form.usuario.data).first():
            flash("El nombre de usuario ya está en uso", "danger")
            return render_template("auth/register.html", form=form)

        # Crear contraseña temporal
        num_doc = form.numero_documento.data
        contrasena_temporal = num_doc[-4:] if len(num_doc) >= 4 else num_doc

        nuevo_empleado = Empleado(
            usuario=form.usuario.data,
            nombre_empleado=form.nombre_empleado.data,
            apellido_empleado=form.apellido_empleado.data,
            numero_documento=form.numero_documento.data,
            tipo_documento=form.tipo_documento.data,
            email=form.email.data,
            telefono=form.telefono.data,
            direccion=form.direccion.data,
            cargo_establecido=form.cargo_establecido.data,
            temporal=True,
            activo=True,
            email_confirmado=False,
            aceptado_terminos=form.aceptar_terminos.data
        )

        nuevo_empleado.set_password(contrasena_temporal)

        
        # 1) Guardar usuario primero para que exista en la BD
        db.session.add(nuevo_empleado)
        db.session.commit()  # ahora tiene id y podemos generar/almacenar el token

        # 2) Generar token y persistirlo (generate_confirmation_token escribe token en el objeto)
        token = nuevo_empleado.generate_confirmation_token()
        db.session.commit()  # persiste token_confirmacion y token_confirmacion_expiry

        # 3) Intentar enviar email; loggear cualquier excepción para ver el error en Vercel
        import traceback
        try:
            enviar_email_confirmacion(nuevo_empleado, token)
            flash(f"¡Registro exitoso! Se ha enviado un email de confirmación a {nuevo_empleado.email}.", "success")
        except Exception as e:
            print("Error enviando email:", e)
            traceback.print_exc()
            flash("Usuario creado, pero hubo un error al enviar el email. Contacte al administrador.", "warning")
        
        # Auditoría
        registrar_auditoria('CREATE', 'empleado', nuevo_empleado.id_empleados, None, {
            'usuario': nuevo_empleado.usuario,
            'nombre': nuevo_empleado.nombre_empleado
        })

        return redirect(url_for("auth.login"))

    return render_template("auth/register.html", form=form)

# NUEVO: Ruta para confirmar email
@auth_bp.route("/confirm/<token>")
def confirm_email(token):
    empleado = Empleado.query.filter_by(token_confirmacion=token).first()
    
    if not empleado:
        flash("Token de confirmación inválido", "danger")
        return redirect(url_for("auth.login"))
    
    if not empleado.verify_confirmation_token(token):
        flash("El token de confirmación ha expirado. Solicita un nuevo email de confirmación.", "danger")
        return redirect(url_for("auth.resend_confirmation"))
    
    empleado.confirmar_email()
    db.session.commit()
    
    registrar_auditoria('CONFIRM_EMAIL', 'empleado', empleado.id_empleados, None, {
        'email_confirmado': True
    })
    
    flash("¡Email confirmado exitosamente! Ya puedes iniciar sesión.", "success")
    return redirect(url_for("auth.login"))

# NUEVO: Reenviar email de confirmación
@auth_bp.route("/resend_confirmation", methods=["GET", "POST"])
def resend_confirmation():
    if request.method == "POST":
        email = request.form.get("email")
        empleado = Empleado.query.filter_by(email=email).first()
        
        if empleado:
            if empleado.email_confirmado:
                flash("Este email ya ha sido confirmado", "info")
                return redirect(url_for("auth.login"))
            
            # Generar nuevo token
            token = empleado.generate_confirmation_token()
            db.session.commit()
            
            try:
                enviar_email_confirmacion(empleado, token)
                flash("Se ha enviado un nuevo email de confirmación", "success")
            except Exception as e:
                flash(f"Error al enviar el email. Contacte al administrador.", "danger")
                print(f"Error: {e}")
        else:
            flash("Si el email existe en nuestro sistema, recibirás un enlace de confirmación", "info")
    
    return render_template("auth/resend_confirmation.html")

@auth_bp.route("/logout")
@login_required
def logout():
    SesionActiva.query.filter_by(
        id_empleados=current_user.id_empleados,
        activa=True
    ).update({'activa': False})
    db.session.commit()
    
    logout_user()
    flash("Sesión cerrada correctamente", "info")
    return redirect(url_for("auth.login"))

@auth_bp.route("/logout_all", methods=["POST"])
@login_required
def logout_all():
    SesionActiva.query.filter_by(id_empleados=current_user.id_empleados).update({'activa': False})
    db.session.commit()
    logout_user()
    flash("Se han cerrado todas las sesiones activas", "success")
    return redirect(url_for("auth.login"))

@auth_bp.route("/change_password", methods=["GET", "POST"])
@login_required
def change_password():
    form = ChangePasswordForm()
    if form.validate_on_submit():
        if current_user.check_password(form.current_password.data):
            current_user.set_password(form.new_password.data)
            current_user.temporal = False
            db.session.commit()
            
            registrar_auditoria('UPDATE', 'empleado', current_user.id_empleados, 
                              {'temporal': True}, {'temporal': False})
            
            flash("Contraseña actualizada exitosamente", "success")
            return redirect(url_for("dashboard.index"))
        else:
            flash("Contraseña actual incorrecta", "danger")
    
    return render_template("auth/change_password.html", form=form)

@auth_bp.route("/request_reset", methods=["GET", "POST"])
def request_password_reset():
    form = RequestPasswordResetForm()
    if form.validate_on_submit():
        empleado = Empleado.query.filter_by(email=form.email.data).first()
        if empleado:
            if not empleado.email_confirmado:
                flash("Primero debe confirmar su email. Revise su correo.", "warning")
                return redirect(url_for("auth.resend_confirmation"))
            
            token = empleado.generate_reset_token()
            db.session.commit()
            
            reset_url = url_for('auth.reset_password', token=token, _external=True)
            msg = Message("Recuperación de Contraseña - Hayuelos",
                        recipients=[empleado.email])
            msg.body = f"""Hola {empleado.nombre_empleado},

Has solicitado restablecer tu contraseña. Haz clic en el siguiente enlace:

{reset_url}

Este enlace expira en 1 hora.

Si no solicitaste este cambio, ignora este correo.

Saludos,
Sistema Hayuelos
"""
            msg.html = f"""
            <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
                    <h2 style="color: #E10000;">Recuperación de Contraseña</h2>
                    <p>Hola <strong>{empleado.nombre_empleado}</strong>,</p>
                    <p>Has solicitado restablecer tu contraseña en Hayuelos.</p>
                    
                    <div style="text-align: center; margin: 30px 0;">
                        <a href="{reset_url}" 
                           style="background-color: #E10000; color: white; padding: 12px 30px; text-decoration: none; border-radius: 5px; display: inline-block;">
                            Restablecer Contraseña
                        </a>
                    </div>
                    
                    <p style="color: #666; font-size: 0.9em;">Este enlace expira en 1 hora.</p>
                    <p style="color: #666; font-size: 0.9em;">Si no solicitaste este cambio, ignora este correo.</p>
                    
                    <hr style="margin: 30px 0; border: none; border-top: 1px solid #ddd;">
                    <p style="color: #999; font-size: 0.8em; text-align: center;">
                        Sistema Hayuelos
                    </p>
                </div>
            </body>
            </html>
            """
            try:
                mail.send(msg)
                flash("Se ha enviado un enlace de recuperación a tu email", "success")
            except Exception as e:
                flash(f"Error al enviar el email. Contacte al administrador.", "danger")
                print(f"Error: {e}")
        else:
            flash("Si el email existe en nuestro sistema, recibirás un enlace de recuperación", "info")
    
    return render_template("auth/request_reset.html", form=form)

@auth_bp.route("/reset/<token>", methods=["GET", "POST"])
def reset_password(token):
    empleado = Empleado.query.filter_by(reset_token=token).first()
    if not empleado or not empleado.verify_reset_token(token):
        flash("Token inválido o expirado", "danger")
        return redirect(url_for("auth.request_password_reset"))
    
    form = PasswordResetForm()  # 👈 El formulario debe estar aquí
    
    if form.validate_on_submit():
        empleado.set_password(form.password.data)
        empleado.reset_token = None
        empleado.reset_token_expiry = None
        empleado.temporal = False
        db.session.commit()
        
        registrar_auditoria('RESET_PASSWORD', 'empleado', empleado.id_empleados, None, {
            'password_reset': True
        })
        
        flash("Contraseña restablecida exitosamente", "success")
        return redirect(url_for("auth.login"))
    
    return render_template("auth/reset_password.html", form=form)  # 👈 Pasando el form

@auth_bp.route("/reset_password/<int:empleado_id>", methods=["POST"])
@login_required
@admin_or_encargado_required
def reset_password_admin(empleado_id):
    empleado = Empleado.query.get_or_404(empleado_id)
    
    contrasena_temporal = empleado.numero_documento[-4:] if len(empleado.numero_documento) >= 4 else empleado.numero_documento
    empleado.set_password(contrasena_temporal)
    empleado.temporal = True
    db.session.commit()

    empleado.temporal = True
    db.session.commit()
    
    registrar_auditoria('UPDATE', 'empleado', empleado_id, None, {'reset_password': True})
    
    flash(f"Contraseña restablecida para {empleado.nombre_empleado}. Nueva contraseña: {contrasena_temporal}", "success")
    return redirect(url_for("dashboard.empleados"))

@auth_bp.route("/resetear-clave/<int:empleado_id>", methods=["POST"])
@login_required
@admin_required
def resetear_clave_empleado(empleado_id):
    empleado = Empleado.query.get_or_404(empleado_id)
    
    temp_password = empleado.numero_documento[-4:] if len(empleado.numero_documento) >= 4 else empleado.numero_documento
    empleado.set_password(temp_password)
    empleado.temporal = True
    db.session.commit()
    
    flash(f'Contraseña reseteada para {empleado.usuario}. Nueva: {temp_password}', 'warning')
    registrar_auditoria('reset_password', 'empleado', empleado.id_empleados, None, {'temporal': True})
    
    return redirect(url_for('dashboard.empleados'))

# ============= MAIN ROUTES =============
@main_bp.route("/")
def home():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard.index"))
    return redirect(url_for("auth.login"))

@main_bp.route("/terminos")
def terminos():
    return render_template("terminos.html")

@main_bp.route("/privacidad")
def privacidad():
    return render_template("privacidad.html")

# ============= DASHBOARD ROUTES =============
@dashboard_bp.route("/")
@login_required
def index():
    tanques = Tanque.query.filter_by(activo=True).all()
    total_capacity = sum(float(t.capacidad) for t in tanques) if tanques else 0
    mediciones_recientes = RegistroMedida.query.order_by(
        RegistroMedida.fecha_hora_registro.desc()
    ).limit(5).all()
    descargues_hoy = Descargue.query.filter_by(fecha=date.today()).all()

    tanques_por_tipo = {}
    for tanque in tanques:
        tipo = tanque.tipo_combustible
        if tipo not in tanques_por_tipo:
            tanques_por_tipo[tipo] = {"count": 0, "capacity": 0, "current": 0}
        tanques_por_tipo[tipo]["count"] += 1
        tanques_por_tipo[tipo]["capacity"] += float(tanque.capacidad)
        tanques_por_tipo[tipo]["current"] += tanque.contenido or 0

    combustible_mas_vendido = db.session.query(
        Tanque.tipo_combustible,
        func.sum(Venta.cantidad_galones).label('total')
    ).join(Venta).group_by(Tanque.tipo_combustible).order_by(func.sum(Venta.cantidad_galones).desc()).first()

    ventas_por_mes = db.session.query(
        extract('month', Venta.fecha).label('mes'),
        func.sum(Venta.cantidad_galones).label('total')
    ).group_by('mes').order_by(func.sum(Venta.cantidad_galones).desc()).all()

    context = {
        "tanques": tanques,
        "total_capacity": total_capacity,
        "mediciones_recientes": mediciones_recientes,
        "descargues_hoy": descargues_hoy,
        "tanques_por_tipo": tanques_por_tipo,
        "total_tanques": len(tanques),
        "combustible_mas_vendido": combustible_mas_vendido,
        "ventas_por_mes": ventas_por_mes
    }
    return render_template("dashboard/index.html", **context)

@dashboard_bp.route("/tanques")
@login_required
def tanques():
    """Mostrar TODOS los tanques (activos e inactivos)"""
    # ❌ ANTES: tanques = Tanque.query.filter_by(activo=True).all()
    # ✅ AHORA: Mostrar todos los tanques, ordenados por activo primero
    
    tanques = Tanque.query.order_by(
        Tanque.activo.desc(),  # Activos primero
        Tanque.id_tanques.asc()  # Luego por ID
    ).all()
    
    return render_template("dashboard/tanques.html", tanques=tanques)

@dashboard_bp.route("/empleados")
@login_required
@admin_or_encargado_required
def empleados():
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    empleados_paginados = Empleado.query.order_by(
        Empleado.cargo_establecido.asc(),
        Empleado.nombre_empleado.asc()
    ).paginate(page=page, per_page=per_page, error_out=False)
    
    return render_template(
        "dashboard/empleados.html", 
        empleados=empleados_paginados.items,
        pagination=empleados_paginados
    )


@dashboard_bp.route("/empleados/cambiar-rol/<int:empleado_id>", methods=["POST"])
@login_required
@admin_or_encargado_required
def cambiar_rol_empleado(empleado_id):
    """Cambiar el rol de un empleado"""
    empleado = Empleado.query.get_or_404(empleado_id)
    nuevo_rol = request.form.get('nuevo_rol')
    
    if nuevo_rol not in ['admin', 'encargado', 'islero']:
        flash("Rol no válido", "danger")
        return redirect(url_for('dashboard.empleados'))
    
    if empleado.id_empleados == current_user.id_empleados:
        flash("No puedes cambiar tu propio rol", "warning")
        return redirect(url_for('dashboard.empleados'))
    
    if current_user.cargo_establecido == 'encargado' and nuevo_rol == 'admin':
        flash("Solo los administradores pueden asignar el rol de administrador", "danger")
        return redirect(url_for('dashboard.empleados'))
    
    rol_anterior = empleado.cargo_establecido
    empleado.cargo_establecido = nuevo_rol
    db.session.commit()
    
    registrar_auditoria('UPDATE', 'empleado', empleado.id_empleados, 
                        {'cargo_establecido': rol_anterior}, 
                        {'cargo_establecido': nuevo_rol})
    
    flash(f"Rol de {empleado.nombre_empleado} cambiado a {nuevo_rol.capitalize()}", "success")
    return redirect(url_for('dashboard.empleados'))

# routes.py - FUNCIÓN estadisticas() CORREGIDA
# Reemplaza SOLO esta función en tu routes.py

@dashboard_bp.route("/estadisticas")
@login_required
@admin_or_encargado_required
def estadisticas():
    """Dashboard de estadísticas para toma de decisiones - 2 estadísticas clave"""
    from collections import OrderedDict
    
    hoy = datetime.now()
    hace_30_dias = hoy - timedelta(days=30)
    
    # ===== ESTADÍSTICA 1: ALERTA DE TANQUES CON STOCK BAJO =====
    tanques = Tanque.query.filter_by(activo=True).all()
    tanques_alerta = []
    
    for tanque in tanques:
        porcentaje = tanque.porcentaje_llenado
        contenido = tanque.contenido
        capacidad = tanque.capacidad or 0
        
        if porcentaje < 30:
            nivel = 'critico' if porcentaje < 15 else 'bajo'
            tanques_alerta.append({
                'id': tanque.id_tanques,
                'tipo': tanque.tipo_combustible,
                'contenido': round(contenido, 2),
                'capacidad': capacidad,
                'porcentaje': round(porcentaje, 1),
                'nivel': nivel,
                'galones_faltantes': round(capacidad - contenido, 2)
            })
    
    tanques_alerta.sort(key=lambda x: x['porcentaje'])
    
    # ===== ESTADÍSTICA 2: RENDIMIENTO Y PÉRDIDAS DEL MES =====
    ventas_por_tipo = OrderedDict()
    cargado_por_tipo = OrderedDict()
    
    for tanque in tanques:
        tipo = tanque.tipo_combustible
        
        mediciones = RegistroMedida.query.filter(
            RegistroMedida.id_tanques == tanque.id_tanques,
            RegistroMedida.fecha_hora_registro >= hace_30_dias
        ).order_by(RegistroMedida.fecha_hora_registro.asc()).all()
        
        total_vendido = 0
        for i in range(1, len(mediciones)):
            try:
                galones_anterior = float(mediciones[i-1].galones or 0)
                galones_actual = float(mediciones[i].galones or 0)
                diferencia = galones_anterior - galones_actual
                if diferencia > 0:
                    total_vendido += diferencia
            except (ValueError, TypeError):
                continue
        
        if total_vendido > 0:
            ventas_por_tipo[tipo] = ventas_por_tipo.get(tipo, 0) + total_vendido
        
        cargues = MedicionCargue.query.filter(
            MedicionCargue.id_tanques == tanque.id_tanques,
            MedicionCargue.fecha >= hace_30_dias
        ).all()
        
        for cargue in cargues:
            try:
                galones = float(cargue.galones_totales or 0)
                cargado_por_tipo[tipo] = cargado_por_tipo.get(tipo, 0) + galones
            except (ValueError, TypeError):
                continue
    
    total_cargado = sum(cargado_por_tipo.values())
    total_vendido = sum(ventas_por_tipo.values())
    diferencia = total_cargado - total_vendido
    porcentaje_eficiencia = (total_vendido / total_cargado * 100) if total_cargado > 0 else 0
    
    rendimiento = {
        'total_cargado': round(total_cargado, 2),
        'total_vendido': round(total_vendido, 2),
        'diferencia': round(diferencia, 2),
        'porcentaje_eficiencia': round(porcentaje_eficiencia, 1),
        'estado': 'normal' if diferencia >= 0 else 'alerta',
        'ventas_por_tipo': dict(ventas_por_tipo),
        'cargado_por_tipo': dict(cargado_por_tipo)
    }
    
    # ===== DATOS PARA REPORTE MENSUAL =====
    ventas_por_dia = OrderedDict()
    mediciones_mes = RegistroMedida.query.filter(
        RegistroMedida.fecha_hora_registro >= hace_30_dias
    ).order_by(RegistroMedida.fecha_hora_registro.asc()).all()
    
    for i in range(1, len(mediciones_mes)):
        try:
            fecha = mediciones_mes[i].fecha_hora_registro.strftime('%Y-%m-%d')
            galones_anterior = float(mediciones_mes[i-1].galones or 0)
            galones_actual = float(mediciones_mes[i].galones or 0)
            diferencia = galones_anterior - galones_actual
            if diferencia > 0:
                ventas_por_dia[fecha] = ventas_por_dia.get(fecha, 0) + diferencia
        except (ValueError, TypeError):
            continue
    
    return render_template(
        'dashboard/estadisticas.html',
        tanques_alerta=tanques_alerta,
        rendimiento=rendimiento,
        ventas_por_tipo=dict(ventas_por_tipo),
        ventas_por_dia=dict(ventas_por_dia),
        fecha_reporte=hoy.strftime('%B %Y')
    )


@dashboard_bp.route("/reporte-mensual")
@login_required
@admin_or_encargado_required
def reporte_mensual():
    """Generar reporte mensual en PDF"""
    from io import BytesIO
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.units import inch
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from collections import OrderedDict
    import base64
    
    hoy = datetime.now()
    hace_30_dias = hoy - timedelta(days=30)
    
    tanques = Tanque.query.filter_by(activo=True).all()
    ventas_por_tipo = OrderedDict()
    ventas_por_dia = OrderedDict()
    
    for tanque in tanques:
        tipo = tanque.tipo_combustible
        mediciones = RegistroMedida.query.filter(
            RegistroMedida.id_tanques == tanque.id_tanques,
            RegistroMedida.fecha_hora_registro >= hace_30_dias
        ).order_by(RegistroMedida.fecha_hora_registro.asc()).all()
        
        for i in range(1, len(mediciones)):
            try:
                fecha = mediciones[i].fecha_hora_registro.strftime('%Y-%m-%d')
                dia_semana = mediciones[i].fecha_hora_registro.strftime('%A')
                galones_anterior = float(mediciones[i-1].galones or 0)
                galones_actual = float(mediciones[i].galones or 0)
                diferencia = galones_anterior - galones_actual
                if diferencia > 0:
                    ventas_por_tipo[tipo] = ventas_por_tipo.get(tipo, 0) + diferencia
                    ventas_por_dia[fecha] = ventas_por_dia.get(fecha, 0) + diferencia
            except (ValueError, TypeError):
                continue
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, textColor=colors.HexColor('#E10000'), alignment=1)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=12, textColor=colors.grey, alignment=1)
    
    elements.append(Paragraph("REPORTE MENSUAL DE COMBUSTIBLE", title_style))
    elements.append(Paragraph(f"Estación Hayuelos - {hoy.strftime('%B %Y')}", subtitle_style))
    elements.append(Spacer(1, 0.3*inch))
    
    if ventas_por_tipo:
        fig1, ax1 = plt.subplots(figsize=(6, 4))
        colores = ['#E10000', '#00E1E1', '#00E100', '#FFA500', '#8A2BE2']
        tipos = list(ventas_por_tipo.keys())
        valores = list(ventas_por_tipo.values())
        ax1.pie(valores, labels=tipos, autopct='%1.1f%%', colors=colores[:len(tipos)], startangle=90)
        ax1.set_title('Combustible Más Vendido por Tipo', fontsize=14, fontweight='bold')
        
        img_buffer1 = BytesIO()
        plt.savefig(img_buffer1, format='png', dpi=150, bbox_inches='tight')
        plt.close(fig1)
        img_buffer1.seek(0)
        
        elements.append(Image(img_buffer1, width=5*inch, height=3.5*inch))
        elements.append(Spacer(1, 0.3*inch))
    
    if ventas_por_dia:
        fig2, ax2 = plt.subplots(figsize=(8, 4))
        fechas = list(ventas_por_dia.keys())[-14:]
        valores_dias = [ventas_por_dia[f] for f in fechas]
        fechas_cortas = [f[-5:] for f in fechas]
        
        barras = ax2.bar(fechas_cortas, valores_dias, color='#E10000', edgecolor='#C20000')
        ax2.set_xlabel('Fecha', fontsize=10)
        ax2.set_ylabel('Galones Vendidos', fontsize=10)
        ax2.set_title('Días con Más Ventas (Últimas 2 Semanas)', fontsize=14, fontweight='bold')
        plt.xticks(rotation=45, ha='right')
        
        for bar, val in zip(barras, valores_dias):
            ax2.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 5, f'{int(val)}', ha='center', va='bottom', fontsize=8)
        
        plt.tight_layout()
        img_buffer2 = BytesIO()
        plt.savefig(img_buffer2, format='png', dpi=150, bbox_inches='tight')
        plt.close(fig2)
        img_buffer2.seek(0)
        
        elements.append(Image(img_buffer2, width=6*inch, height=3*inch))
        elements.append(Spacer(1, 0.3*inch))
    
    elements.append(Paragraph("RESUMEN DE CONSUMO POR TIPO", styles['Heading2']))
    nota_style = ParagraphStyle('Nota', parent=styles['Normal'], fontSize=9, textColor=colors.grey, alignment=0)
    elements.append(Paragraph("* Los datos de consumo se calculan a partir de las diferencias en las mediciones de inventario.", nota_style))
    elements.append(Spacer(1, 0.1*inch))
    if ventas_por_tipo:
        data = [['Tipo de Combustible', 'Galones Vendidos', 'Porcentaje']]
        total = sum(ventas_por_tipo.values())
        for tipo, galones in ventas_por_tipo.items():
            porcentaje = (galones / total * 100) if total > 0 else 0
            data.append([tipo, f"{galones:,.0f}", f"{porcentaje:.1f}%"])
        data.append(['TOTAL', f"{total:,.0f}", '100%'])
        
        table = Table(data, colWidths=[2.5*inch, 2*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E10000')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFE0E0')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
    
    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph(f"Generado el {hoy.strftime('%d/%m/%Y a las %H:%M')}", subtitle_style))
    
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'reporte_mensual_{hoy.strftime("%Y_%m")}.pdf',
        mimetype='application/pdf'
    )


# ===== RUTAS PARA MANUALES Y POLÍTICAS =====

@main_bp.route("/manual-usuario")
@login_required
def manual_usuario():
    """Manual de usuario (pendiente de subir)"""
    return render_template("documentacion/manual_usuario.html")


@main_bp.route("/manual-tecnico")
@login_required
def manual_tecnico():
    """Manual técnico (pendiente de subir)"""
    return render_template("documentacion/manual_tecnico.html")


@main_bp.route("/politicas")
@login_required
def politicas():
    """Políticas de la empresa (pendiente de subir)"""
    return render_template("documentacion/politicas.html")

# ============= MEDICION ROUTES =============
@medicion_bp.route("/registro", methods=["GET", "POST"])
@login_required
@islero_or_encargado_required
def registro():
    form = MedicionForm()
    tanques = Tanque.query.filter_by(activo=True).all()
    form.tanque.choices = [(t.id_tanques, f"{t.tipo_combustible} - Tanque {t.id_tanques}") for t in tanques]
    
    if form.validate_on_submit():
        tanque_seleccionado = Tanque.query.get(form.tanque.data)
        
        # Fix: Compute altura_maxima_cm if None
        if tanque_seleccionado.altura_maxima_cm is None:
            tanque_seleccionado.altura_maxima_cm = calcular_altura_maxima(tanque_seleccionado.capacidad)
            db.session.commit()  # Save the update to DB
        
        medida_str = form.medida_combustible.data.replace(',', '.')
        medida_cm = float(medida_str)
        
        # Fix: Only compare if altura_maxima_cm is not None (though we just set it)
        if tanque_seleccionado.altura_maxima_cm is not None and medida_cm > tanque_seleccionado.altura_maxima_cm:
            flash(
                f"❌ La medida ({medida_cm} cm) supera la altura máxima del tanque ({tanque_seleccionado.altura_maxima_cm} cm)", 
                "danger"
            )
            return render_template("medicion/registro.html", form=form)
        
        # Rest of the code remains the same...
    if form.validate_on_submit():
        imagen_path = None
        if form.imagen.data:
            file = form.imagen.data
            if allowed_file(file.filename):
                filename = secure_filename(f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                filepath = os.path.join('static/uploads', filename)
                os.makedirs('static/uploads', exist_ok=True)
                file.save(filepath)
                imagen_path = filename

        medicion = RegistroMedida(
            medida_combustible=form.medida_combustible.data,
            id_empleados=current_user.id_empleados,
            fecha_hora_registro=datetime.now(),
            galones=form.galones.data,
            id_tanques=form.tanque.data,
            tipo_medida=form.tipo_medida.data,
            novedad=form.novedad.data,
            imagen_path=imagen_path
        )
        db.session.add(medicion)
        db.session.commit()
        
        registrar_auditoria('CREATE', 'registro_medidas', medicion.id_registro_medidas, None, {
            'tanque': form.tanque.data,
            'galones': form.galones.data
        })

        flash("Medición registrada exitosamente", "success")
        return redirect(url_for("medicion.historial"))

    return render_template("medicion/registro.html", form=form)

@medicion_bp.route("/historial")
@login_required
def historial():
    page = request.args.get("page", 1, type=int)
    
    query = RegistroMedida.query
    
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    tanque_id = request.args.get('tanque', type=int)
    tipo = request.args.get('tipo')
    
    if fecha_desde:
        query = query.filter(RegistroMedida.fecha_hora_registro >= fecha_desde)
    if fecha_hasta:
        query = query.filter(RegistroMedida.fecha_hora_registro <= fecha_hasta)
    if tanque_id:
        query = query.filter_by(id_tanques=tanque_id)
    if tipo:
        query = query.filter_by(tipo_medida=tipo)
    
    mediciones = query.order_by(
        RegistroMedida.fecha_hora_registro.desc()
    ).paginate(page=page, per_page=20, error_out=False)
    
    tanques = Tanque.query.filter_by(activo=True).all()
    return render_template("medicion/historial.html", mediciones=mediciones, tanques=tanques)

@medicion_bp.route("/descargue", methods=["GET", "POST"])
@login_required
@islero_or_encargado_required
def descargue():
    form = DescargueForm()

    # Cargar tanques activos
    tanques = Tanque.query.filter_by(activo=True).all()
    form.tanque.choices = [(t.id_tanques, f"{t.tipo_combustible} - Tanque {t.id_tanques}") for t in tanques]

    if form.validate_on_submit():
        tanque_seleccionado = Tanque.query.get(form.tanque.data)

        # CORREGIDO: Calcular altura máxima si está vacía
        if tanque_seleccionado.altura_maxima_cm is None:
            tanque_seleccionado.altura_maxima_cm = calcular_altura_maxima(tanque_seleccionado.capacidad)
            db.session.commit()

        try:
            medida_inicial_cm = float(str(form.medida_inicial_cm.data).replace(',', '.'))
            descargue_cm_valor = float(str(form.descargue_cm.data).replace(',', '.'))
            medida_final_cm = float(str(form.medida_final_cm.data).replace(',', '.'))
        except (ValueError, AttributeError) as e:
            flash("Error en los valores de medidas (cm). Verifica que sean números válidos.", "danger")
            return render_template("medicion/descargue.html", form=form)

        # Validaciones de altura máxima
        if medida_inicial_cm > tanque_seleccionado.altura_maxima_cm:
            flash(f"Medida inicial ({medida_inicial_cm} cm) supera la altura máxima del tanque ({tanque_seleccionado.altura_maxima_cm:.2f} cm)", "danger")
            return render_template("medicion/descargue.html", form=form)

        if medida_final_cm > tanque_seleccionado.altura_maxima_cm:
            flash(f"Medida final ({medida_final_cm} cm) supera la altura máxima del tanque ({tanque_seleccionado.altura_maxima_cm:.2f} cm)", "danger")
            return render_template("medicion/descargue.html", form=form)

        # Guardar imagen si existe
        imagen_path = None
        if form.imagen.data:
            file = form.imagen.data
            if allowed_file(file.filename):
                filename = secure_filename(f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                filepath = os.path.join('static/uploads', filename)
                os.makedirs('static/uploads', exist_ok=True)
                file.save(filepath)
                imagen_path = filename

        # Crear el descargue
        descargue_obj = Descargue(
            id_empleados=current_user.id_empleados,
            medida_inicial_cm=form.medida_inicial_cm.data,
            medida_inicial_gl=form.medida_inicial_gl.data,
            descargue_cm=form.descargue_cm.data,
            descargue_gl=form.descargue_gl.data,
            medida_final_cm=form.medida_final_cm.data,
            medida_final_gl=form.medida_final_gl.data,
            diferencia=form.diferencia.data,
            tanque=form.tanque.data,
            observaciones1=form.observaciones1.data,
            observaciones2=form.observaciones2.data,
            kit_derrames='si' if form.kit_derrames.data else 'no',
            extintores='si' if form.extintores.data else 'no',
            conos='si' if form.conos.data else 'no',
            boquillas='si' if form.boquillas.data else 'no',
            botas='si' if form.botas.data else 'no',
            gafas='si' if form.gafas.data else 'no',
            tapaoidos='si' if form.tapaoidos.data else 'no',
            guantes='si' if form.guantes.data else 'no',
            brillante=form.brillante.data,
            traslucido=form.traslucido.data,
            claro=form.claro.data,
            solidos=form.solidos.data,
            separacion=form.separacion.data,
            fecha=form.fecha.data or date.today(),
            imagen_path=imagen_path
        )
        db.session.add(descargue_obj)
        db.session.commit()

        registrar_auditoria('CREATE', 'descargues', descargue_obj.idDescargue, None, {
            'tanque': form.tanque.data,
            'galones_descargados': form.diferencia.data
        })

        flash("Descargue registrado exitosamente", "success")
        return redirect(url_for("medicion.historial_descargues"))

    return render_template("medicion/descargue.html", form=form)

@medicion_bp.route("/api/convert_cm_to_gallons/<int:tanque_id>", methods=["GET"])
@login_required
def convert_cm_to_gallons(tanque_id):
    """API para convertir cm a galones según el tanque específico"""
    cm = request.args.get('cm', type=float, default=0)
    
    tanque = Tanque.query.get_or_404(tanque_id)
    
    # Fix: Use tanque.radio_cm (from DB) instead of non-existent diametro_m
    radio_cm = tanque.radio_cm if tanque.radio_cm else 125  # Fallback to default if None
    
    # Fórmula correcta: V = π * r² * h
    # 1 galón = 3785.411784 cm³
    area_base = 3.14159 * (radio_cm ** 2)  # cm²
    volumen_cm3 = area_base * cm  # cm³
    galones = volumen_cm3 / 3785.411784
    
    return jsonify({
        'cm': cm,
        'gallons': round(galones, 2),
        'tanque_id': tanque_id,
        'radio_cm': radio_cm
    })

@medicion_bp.route("/historial_descargues")
@login_required
def historial_descargues():
    page = request.args.get("page", 1, type=int)
    descargues = Descargue.query.order_by(Descargue.fecha.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    return render_template("medicion/historial_descargues.html", descargues=descargues)

# ============= ADMIN ROUTES =============
@admin_bp.route("/toggle_empleado/<int:empleado_id>", methods=["POST"])
@login_required
@admin_required
def toggle_empleado(empleado_id):
    empleado = Empleado.query.get_or_404(empleado_id)
    empleado.activo = not empleado.activo
    db.session.commit()
    
    registrar_auditoria('UPDATE', 'empleado', empleado_id, 
                      {'activo': not empleado.activo}, {'activo': empleado.activo})
    
    estado = "habilitado" if empleado.activo else "deshabilitado"
    flash(f"Empleado {empleado.nombre_empleado} {estado}", "success")
    return redirect(url_for("dashboard.empleados"))

@admin_bp.route("/carga_masiva", methods=["GET", "POST"])
@login_required
@admin_required
def carga_masiva():
    form = CargaMasivaForm()
    if form.validate_on_submit():
        file = form.archivo.data
        tipo_carga = form.tipo_carga.data
        
        if not file or not allowed_file(file.filename, {'csv', 'xlsx', 'xls'}):
            flash("Archivo inválido", "danger")
            return redirect(request.url)

        try:
            if file.filename.lower().endswith('.csv'):
                df = pd.read_csv(file, sep=',', decimal='.', encoding='utf-8-sig')
            else:
                df = pd.read_excel(file)
            
            df = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
            df.replace({"True": True, "False": False}, inplace=True, regex=True)

            count = 0
            errors = []

            if tipo_carga == 'empleados':
                required = ['nombre_empleado', 'apellido_empleado', 'numero_documento', 'email', 'usuario']
                if not all(col in df.columns for col in required):
                    flash(f"Faltan columnas: {', '.join(required)}", "danger")
                    return redirect(request.url)
                
                for idx, row in df.iterrows():
                    if Empleado.query.filter_by(numero_documento=row['numero_documento']).first():
                        errors.append(f"Fila {idx+2}: Documento duplicado")
                        continue
                    if Empleado.query.filter_by(usuario=row['usuario']).first():
                        errors.append(f"Fila {idx+2}: Usuario duplicado")
                        continue
                    
                    temp_pass = str(row['numero_documento'])[-4:]
                    empleado = Empleado(
                        nombre_empleado=row['nombre_empleado'],
                        apellido_empleado=row['apellido_empleado'],
                        numero_documento=row['numero_documento'],
                        tipo_documento=row.get('tipo_documento', 'CC'),
                        email=row['email'],
                        telefono=row.get('telefono', ''),
                        direccion=row.get('direccion', ''),
                        cargo_establecido=row.get('cargo_establecido', 'Islero'),
                        usuario=row['usuario'],
                        temporal=True,
                        activo=row.get('activo', True),
                        email_confirmado=True,
                        aceptado_terminos=row.get('aceptado_terminos', False)
                    )
                    empleado.set_password(temp_pass)

                    db.session.add(empleado)
                    count += 1

            elif tipo_carga == 'tanques':
                required = ['tipo_combustible', 'capacidad']
                if not all(col in df.columns for col in required):
                    flash("Faltan columnas: tipo_combustible, capacidad", "danger")
                    return redirect(request.url)
                
                for idx, row in df.iterrows():
                    try:
                        capacidad = int(float(str(row['capacidad']).replace(',', '.')))
                        tanque = Tanque(
                            tipo_combustible=row['tipo_combustible'],
                            capacidad=capacidad,
                            activo=row.get('activo', True)
                        )
                        db.session.add(tanque)
                        count += 1
                    except Exception as e:
                        errors.append(f"Fila {idx+2}: Capacidad inválida → {str(e)}")

            elif tipo_carga == 'mediciones':
                required = ['tanque_id', 'medida_combustible', 'galones', 'tipo_medida', 'fecha_hora_registro', 'empleado_id']
                if not all(col in df.columns for col in required):
                    flash("Faltan columnas en mediciones", "danger")
                    return redirect(request.url)
                
                for idx, row in df.iterrows():
                    try:
                        tanque_id = int(row['tanque_id'])
                        empleado_id = int(row['empleado_id'])
                        tanque = Tanque.query.get(tanque_id)
                        empleado = Empleado.query.get(empleado_id)
                        
                        if not tanque:
                            errors.append(f"Fila {idx+2}: tanque_id {tanque_id} no existe")
                            continue
                        if not empleado:
                            errors.append(f"Fila {idx+2}: empleado_id {empleado_id} no existe")
                            continue
                        
                        medida_str = str(row['medida_combustible']).replace(',', '.')
                        galones_str = str(row['galones']).replace(',', '.')
                        medida = float(medida_str)
                        galones = float(galones_str)
                        
                        fecha_str = str(row['fecha_hora_registro']).strip()
                        try:
                            fecha = datetime.strptime(fecha_str, '%Y-%m-%d %H:%M:%S')
                        except:
                            fecha = datetime.strptime(fecha_str, '%Y-%m-%d %H:%M')
                        
                        medicion = RegistroMedida(
                            id_tanques=tanque.id_tanques,
                            id_empleados=empleado.id_empleados,
                            medida_combustible=medida,
                            galones=galones,
                            tipo_medida=row['tipo_medida'],
                            novedad=row.get('novedad', ''),
                            fecha_hora_registro=fecha
                        )
                        db.session.add(medicion)
                        count += 1
                    except Exception as e:
                        errors.append(f"Fila {idx+2}: Error → {str(e)}")

            db.session.commit()
            registrar_auditoria('CREATE_BULK', tipo_carga, None, None, {'count': count})

            msg = f"Se cargaron {count} registros exitosamente"
            if errors:
                error_sample = "; ".join(errors[:3])
                flash(f"{msg}. Errores: {len(errors)} → {error_sample}", "warning")
            else:
                flash(msg, "success")

        except Exception as e:
            db.session.rollback()
            flash(f"Error crítico: {str(e)}", "danger")
            print(f"[ERROR] {e}")

    return render_template("admin/carga_masiva.html", form=form)


# ============= EXPORT ROUTES (AGREGAR AL FINAL DE routes.py) =============

@admin_bp.route("/export_menu", methods=["GET"])
@login_required
@admin_or_encargado_required
def export_menu():
    """Menú de exportación de datos"""
    tanques = Tanque.query.filter_by(activo=True).all()
    return render_template("admin/export_menu.html", tanques=tanques)


@admin_bp.route("/export/<tipo>", methods=["GET"])
@login_required
@admin_or_encargado_required
def export_data(tipo):
    """Exportar datos en formato Excel o CSV"""
    from io import BytesIO
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash("Error: Instale openpyxl con 'pip install openpyxl'", "danger")
        return redirect(url_for('dashboard.index'))
    
    formato = request.args.get('formato', 'excel')  # 'excel' o 'csv'
    
    # Determinar qué datos exportar
    if tipo == 'empleados':
        empleados = Empleado.query.all()
        data = []
        headers = ['ID', 'Usuario', 'Nombre', 'Apellido', 'Documento', 'Email', 'Teléfono', 
                   'Dirección', 'Cargo', 'Activo', 'Email Confirmado', 'Fecha Creación']
        
        for emp in empleados:
            data.append([
                emp.id_empleados,
                emp.usuario,
                emp.nombre_empleado,
                emp.apellido_empleado,
                emp.numero_documento,
                emp.email,
                emp.telefono or '',
                emp.direccion or '',
                emp.cargo_establecido,
                'Sí' if emp.activo else 'No',
                'Sí' if emp.email_confirmado else 'No',
                emp.fecha_creacion.strftime('%Y-%m-%d %H:%M') if emp.fecha_creacion else ''
            ])
        filename = f"empleados_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
    elif tipo == 'tanques':
        tanques = Tanque.query.all()
        data = []
        headers = ['ID', 'Tipo Combustible', 'Capacidad (gal)', 'Contenido (gal)', 
                   'Volumen (m³)', 'Activo', 'Fecha Creación']
        
        for tanque in tanques:
            data.append([
                tanque.id_tanques,
                tanque.tipo_combustible,
                tanque.capacidad,
                tanque.contenido or 0,
                tanque.volumen_m3,
                'Sí' if tanque.activo else 'No',
                tanque.fecha_creacion.strftime('%Y-%m-%d') if tanque.fecha_creacion else ''
            ])
        filename = f"tanques_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
    elif tipo == 'mediciones':
        # Filtros opcionales
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        tanque_id = request.args.get('tanque_id')
        
        query = RegistroMedida.query
        if fecha_desde:
            query = query.filter(RegistroMedida.fecha_hora_registro >= fecha_desde)
        if fecha_hasta:
            query = query.filter(RegistroMedida.fecha_hora_registro <= fecha_hasta)
        if tanque_id:
            query = query.filter_by(id_tanques=int(tanque_id))
        
        mediciones = query.order_by(RegistroMedida.fecha_hora_registro.desc()).all()
        
        data = []
        headers = ['ID', 'Fecha/Hora', 'Tanque', 'Tipo Combustible', 'Medida (cm)', 
                   'Galones', 'Tipo Medición', 'Empleado', 'Novedad']
        
        for med in mediciones:
            data.append([
                med.id_registro_medidas,
                med.fecha_hora_registro.strftime('%Y-%m-%d %H:%M:%S') if med.fecha_hora_registro else '',
                f"Tanque {med.tanque.id_tanques}" if med.tanque else 'N/A',
                med.tanque.tipo_combustible if med.tanque else 'N/A',
                med.medida_combustible or '',
                med.galones or 0,
                med.tipo_medida or 'rutinario',
                med.empleado.nombre_empleado if med.empleado else 'N/A',
                med.novedad or ''
            ])
        filename = f"mediciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
    elif tipo == 'descargues':
        # Filtros opcionales
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        
        query = Descargue.query
        if fecha_desde:
            query = query.filter(Descargue.fecha >= fecha_desde)
        if fecha_hasta:
            query = query.filter(Descargue.fecha <= fecha_hasta)
        
        descargues = query.order_by(Descargue.fecha.desc()).all()
        
        data = []
        headers = ['ID', 'Fecha', 'Tanque', 'Medida Inicial (gl)', 'Descargue (gl)', 
                   'Medida Final (gl)', 'Diferencia', 'Empleado', 'Kit Derrames', 
                   'Extintores', 'Observaciones']
        
        for desc in descargues:
            data.append([
                desc.idDescargue,
                desc.fecha.strftime('%Y-%m-%d') if desc.fecha else '',
                desc.tanque or '',
                float(desc.medida_inicial_gl) if desc.medida_inicial_gl else 0,
                float(desc.descargue_gl) if desc.descargue_gl else 0,
                float(desc.medida_final_gl) if desc.medida_final_gl else 0,
                float(desc.diferencia) if desc.diferencia else 0,
                desc.empleado.nombre_empleado if desc.empleado else 'N/A',
                desc.kit_derrames or 'no',
                desc.extintores or 'no',
                desc.observaciones1 or ''
            ])
        filename = f"descargues_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    else:
        flash("Tipo de exportación no válido", "danger")
        return redirect(url_for('dashboard.index'))
    
    # Generar archivo según formato
    if formato == 'csv':
        # Exportar como CSV
        output = BytesIO()
        df = pd.DataFrame(data, columns=headers)
        df.to_csv(output, index=False, encoding='utf-8-sig')
        output.seek(0)
        
        registrar_auditoria('EXPORT_CSV', tipo, None, None, {
            'formato': 'csv',
            'registros': len(data)
        })
        
        return send_file(
            output,
            mimetype='text/csv',
            as_attachment=True,
            download_name=f"{filename}.csv"
        )
    
    else:  # Excel
        # Crear workbook de Excel con estilos
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = tipo.capitalize()
        
        # Estilo para encabezados
        header_fill = PatternFill(start_color="E10000", end_color="E10000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Escribir datos
        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        registrar_auditoria('EXPORT_EXCEL', tipo, None, None, {
            'formato': 'excel',
            'registros': len(data)
        })
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{filename}.xlsx"
        )

# Agregar estas rutas al final de admin_bp en routes.py

# ============= GESTIÃ"N COMPLETA DE TANQUES =============

# ============= GESTIÓN COMPLETA DE TANQUES =============

@admin_bp.route("/tanques/crear", methods=["GET", "POST"])
@login_required
@admin_or_encargado_required
def crear_tanque():
    """Crear nuevo tanque"""
    form = TanqueForm()
    if form.validate_on_submit():
        tanque = Tanque(
            tipo_combustible=form.tipo_combustible.data,
            capacidad=form.capacidad.data,
            activo=form.activo.data,
            altura_maxima_cm=calcular_altura_maxima(form.capacidad.data),
            radio_cm=125.0
        )
        db.session.add(tanque)
        db.session.commit()
        
        registrar_auditoria('CREATE', 'tanques', tanque.id_tanques, None, {
            'tipo': form.tipo_combustible.data,
            'capacidad': form.capacidad.data
        })
        
        flash("Tanque creado exitosamente", "success")
        return redirect(url_for("dashboard.tanques"))
    
    return render_template("admin/tanque_form.html", form=form, titulo="Nuevo Tanque", accion="crear")


@admin_bp.route("/tanques/<int:tanque_id>/editar", methods=["GET", "POST"])
@login_required
@admin_or_encargado_required
def editar_tanque(tanque_id):
    """Editar tanque existente"""
    tanque = Tanque.query.get_or_404(tanque_id)
    form = TanqueForm(obj=tanque)
    
    if form.validate_on_submit():
        datos_anteriores = {
            'tipo': tanque.tipo_combustible,
            'capacidad': tanque.capacidad,
            'activo': tanque.activo
        }
        
        tanque.tipo_combustible = form.tipo_combustible.data
        tanque.capacidad = form.capacidad.data
        tanque.activo = form.activo.data
        tanque.altura_maxima_cm = calcular_altura_maxima(form.capacidad.data)
        
        db.session.commit()
        
        registrar_auditoria('UPDATE', 'tanques', tanque_id, datos_anteriores, {
            'tipo': tanque.tipo_combustible,
            'capacidad': tanque.capacidad,
            'activo': tanque.activo
        })
        
        flash("Tanque actualizado exitosamente", "success")
        return redirect(url_for("dashboard.tanques"))
    
    return render_template("admin/tanque_form.html", form=form, titulo="Editar Tanque", accion="editar", tanque=tanque)


@admin_bp.route("/tanques/<int:tanque_id>/toggle", methods=["POST"])
@login_required
@admin_or_encargado_required
def toggle_tanque(tanque_id):
    """Activar/Desactivar tanque"""
    tanque = Tanque.query.get_or_404(tanque_id)
    tanque.activo = not tanque.activo
    db.session.commit()
    
    registrar_auditoria('UPDATE', 'tanques', tanque_id, 
                      {'activo': not tanque.activo}, {'activo': tanque.activo})
    
    estado = "activado" if tanque.activo else "desactivado"
    flash(f"Tanque {estado}", "success")
    return redirect(url_for("dashboard.tanques"))


def calcular_altura_maxima(capacidad_galones):
    """Calcular altura máxima en cm basada en capacidad del tanque"""
    # Radio estándar en cm (ajustar según tanques reales)
    radio_cm = 125  # 2.5m de diámetro
    
    # Volumen en cm³ = capacidad en galones * 3785.411784
    volumen_cm3 = capacidad_galones * 3785.411784
    
    # Altura = Volumen / (π * r²)
    area_base = 3.14159 * (radio_cm ** 2)
    altura_cm = volumen_cm3 / area_base
    
    return round(altura_cm, 2)

# ============= CARGUE DE EMERGENCIA =============
@medicion_bp.route("/cargue_emergencia", methods=["GET", "POST"])
@login_required
@islero_or_encargado_required
def cargue_emergencia():
    from forms import CargueEmergenciaForm
    form = CargueEmergenciaForm()
    
    tanques = Tanque.query.filter_by(activo=True).all()
    form.tanque.choices = [(t.id_tanques, f"{t.tipo_combustible} - Tanque {t.id_tanques}") for t in tanques]
    
    if form.validate_on_submit():
        tanque_seleccionado = Tanque.query.get(form.tanque.data)
        
        # Fix: Compute altura_maxima_cm if None
        if tanque_seleccionado.altura_maxima_cm is None:
            tanque_seleccionado.altura_maxima_cm = calcular_altura_maxima(tanque_seleccionado.capacidad)
            db.session.commit()
        
        medida_anterior = float(form.medida_anterior.data.replace(',', '.'))
        medida_posterior = float(form.medida_posterior.data.replace(',', '.'))
        
        # Fix: Only compare if altura_maxima_cm is not None
        if tanque_seleccionado.altura_maxima_cm is not None:
            if medida_anterior > tanque_seleccionado.altura_maxima_cm:
                flash(
                    f"❌ Medida anterior ({medida_anterior} cm) supera altura máxima ({tanque_seleccionado.altura_maxima_cm} cm)", 
                    "danger"
                )
                return render_template("medicion/cargue_emergencia.html", form=form)
            
            if medida_posterior > tanque_seleccionado.altura_maxima_cm:
                flash(
                    f"❌ Medida posterior ({medida_posterior} cm) supera altura máxima ({tanque_seleccionado.altura_maxima_cm} cm)", 
                    "danger"
                )
                return render_template("medicion/cargue_emergencia.html", form=form)
        
        # Rest of the code remains the same...
    
    # Cargar tanques en el selector
    tanques = Tanque.query.filter_by(activo=True).all()
    form.tanque.choices = [(t.id_tanques, f"{t.tipo_combustible} - Tanque {t.id_tanques}") for t in tanques]
    
    if form.validate_on_submit():
        imagen_path = None
        if form.imagen.data:
            file = form.imagen.data
            if allowed_file(file.filename):
                filename = secure_filename(f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                filepath = os.path.join('static/uploads', filename)
                os.makedirs('static/uploads', exist_ok=True)
                file.save(filepath)
                imagen_path = filename
        
        cargue = MedicionCargue(
            id_empleados=current_user.id_empleados,
            id_tanques=form.tanque.data,
            medida_anterior=form.medida_anterior.data,
            medida_posterior=form.medida_posterior.data,
            formato_de_entrega=form.formato_entrega.data,
            galones_totales=form.galones_totales.data,
            fecha=datetime.now()
        )
        db.session.add(cargue)
        db.session.commit()
        
        registrar_auditoria('CREATE', 'cargue_emergencia', cargue.id_medicion_cargue, None, {
            'tanque': form.tanque.data,
            'galones': form.galones_totales.data
        })
        
        flash("Cargue de emergencia registrado exitosamente", "success")
        return redirect(url_for("medicion.historial_cargues"))
    
    return render_template("medicion/cargue_emergencia.html", form=form)


@medicion_bp.route("/historial_cargues")
@login_required
@islero_or_encargado_required
def historial_cargues():
    """Historial de cargues de emergencia"""
    page = request.args.get("page", 1, type=int)
    cargues = MedicionCargue.query.order_by(
        MedicionCargue.fecha.desc()
    ).paginate(page=page, per_page=20, error_out=False)
    
    return render_template("medicion/historial_cargues.html", cargues=cargues)

@admin_bp.route("/update_alturas")
@login_required
@admin_required
def update_alturas():
    tanques = Tanque.query.all()
    for t in tanques:
        if t.altura_maxima_cm is None:
            t.altura_maxima_cm = calcular_altura_maxima(t.capacidad)
    db.session.commit()
    flash("Alturas máximas actualizadas para todos los tanques", "success")
    return redirect(url_for("dashboard.tanques"))
